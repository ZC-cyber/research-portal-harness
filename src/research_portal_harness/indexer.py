from __future__ import annotations

from pathlib import Path
from typing import Any

from .common import append_jsonl, now_iso, read_jsonl, sha256_file
from .manifest import load_manifest, manifest_path, unique_records
from .workspace import load_brokers


def _pdf_summary(path: Path) -> dict[str, Any]:
    try:
        import fitz
    except ImportError:
        return {"extract_status": "pymupdf_missing"}
    try:
        doc = fitz.open(path)
        text = ""
        for page in doc[: min(3, len(doc))]:
            text += page.get_text("text")[:2000]
        return {"pages": len(doc), "text_preview": text[:1200], "extract_status": "ok"}
    except Exception as exc:
        return {"extract_status": "error", "error": str(exc)}


def _excel_summary(path: Path) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError:
        return {"extract_status": "openpyxl_missing"}
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
        return {
            "sheets": workbook.sheetnames[:30],
            "sheet_count": len(workbook.sheetnames),
            "extract_status": "ok",
        }
    except Exception as exc:
        return {"extract_status": "error", "error": str(exc)}


def _csv_summary(path: Path) -> dict[str, Any]:
    try:
        with path.open("r", encoding="utf-8", errors="ignore") as handle:
            lines = [next(handle).strip() for _ in range(5)]
        return {"preview_lines": lines, "extract_status": "ok"}
    except StopIteration:
        return {"preview_lines": [], "extract_status": "ok"}
    except Exception as exc:
        return {"extract_status": "error", "error": str(exc)}


def summarize_file(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return _pdf_summary(path)
    if suffix in {".xls", ".xlsx", ".xlsm"}:
        return _excel_summary(path)
    if suffix in {".csv", ".json"}:
        return _csv_summary(path)
    return {"extract_status": "unsupported"}


def index_downloads(root: Path, task_id: str | None = None) -> Path:
    config = load_brokers(root)
    manifest = load_manifest(manifest_path(root, config))
    records = unique_records(manifest)
    if task_id:
        records = [record for record in records if f"/{task_id}/" in record.get("local_path", "")]

    index_path = root / "data" / "indexes" / f"{task_id or 'all'}.documents.jsonl"
    if index_path.exists():
        index_path.unlink()

    rows: list[dict[str, Any]] = []
    for record in records:
        path = Path(record["local_path"])
        if not path.exists():
            continue
        summary = summarize_file(path)
        rows.append(
            {
                "indexed_at": now_iso(),
                "task_id": task_id,
                "broker_id": record.get("broker_id"),
                "broker": record.get("broker"),
                "title": record.get("title"),
                "source_url": record.get("source_url"),
                "local_path": str(path),
                "sha256": record.get("sha256") or sha256_file(path),
                "size_bytes": path.stat().st_size,
                "file_type": path.suffix.lower().lstrip("."),
                "summary": summary,
            }
        )
    append_jsonl(index_path, rows)
    return index_path


def status(root: Path, task_id: str | None = None) -> dict[str, Any]:
    config = load_brokers(root)
    manifest = load_manifest(manifest_path(root, config))
    records = unique_records(manifest)
    if task_id:
        records = [record for record in records if f"/{task_id}/" in record.get("local_path", "")]
    index_path = root / "data" / "indexes" / f"{task_id or 'all'}.documents.jsonl"
    indexed = read_jsonl(index_path)
    return {
        "task_id": task_id,
        "downloaded_files": len(records),
        "indexed_documents": len(indexed),
        "manifest_path": str(manifest_path(root, config)),
        "index_path": str(index_path),
        "by_broker": {
            broker: len([record for record in records if record.get("broker_id") == broker])
            for broker in sorted({record.get("broker_id") for record in records if record.get("broker_id")})
        },
    }

